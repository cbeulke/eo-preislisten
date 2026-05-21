"""
Preisvergleich-App für Photovoltaik-Produkte
Lieferanten: TWE Solar GmbH, Pesasun GmbH, pv partners AG
Formate: PDF, erweiterbar auf CSV/JSON
"""

import io
import re
import secrets
from pathlib import Path

import pandas as pd
import pdfplumber
import streamlit as st
import streamlit_authenticator as stauth
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from streamlit_authenticator.utilities import Hasher

# ── Konstanten ────────────────────────────────────────────────────────────────

ANBIETER_TWE = "TWE"
ANBIETER_PESASUN = "Pesasun"
ANBIETER_PVPARTNERS = "pv partners"
ANBIETER_OEKOTEAM = "Ökoteam Solar"

# Kategorien
KAT_MODULE = "Solarmodule"
KAT_SPEICHER = "Speichersysteme"
KAT_WECHSELRICHTER = "Wechselrichter"

PREISTYP_CONTAINER = "Container"
PREISTYP_PALETTE = "Palette"
PREISTYP_STUECK = "Stück"

SCHEMA = [
    "Anbieter", "Kategorie", "Produkt", "Preistyp",
    "Preis", "Einheit", "Datum", "Quelle",
]

FOOTER_KEYWORDS = [
    "TWE Solar", "Oberer Stadtwald", "Haarbachweg",
    "94486", "94474", "Tel.:", "Web.:", "E-Mail:",
]

# ── Hilfsfunktionen ───────────────────────────────────────────────────────────

# Handles German thousand-separator format (1.234,56) AND compact 4+ digit format (1230)
_GERMAN_PRICE_RE = re.compile(
    r"(\d{1,3}(?:\.\d{3})+(?:,\d+)?|\d{1,3}(?:,\d+)?|\d{4,}(?:,\d+)?)\s*€"
)


def _parse_eur(text: str) -> float | None:
    """'1.234,56 €'  →  1234.56, None wenn nicht parsbar."""
    s = text.strip().replace("€", "").replace(" ", "")
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _all_prices_in_line(line: str) -> list[float]:
    """Alle EUR-Preise aus einer Zeile extrahieren."""
    return [_parse_eur(m) for m in _GERMAN_PRICE_RE.findall(line)
            if _parse_eur(m) is not None]


def _is_footer(line: str) -> bool:
    return any(kw in line for kw in FOOTER_KEYWORDS)


def _detect_date_from_filename(filename: str) -> str:
    """'..._01.05.pdf' oder '..._01.05..pdf' → '2026-05-01'."""
    m = re.search(r"_(\d{2})\.(\d{2})\.?(?:\.pdf)?$", filename, re.IGNORECASE)
    if m:
        day, month = m.group(1), m.group(2)
        return f"2026-{month}-{day}"
    return "unbekannt"


def _detect_twe_type(filename: str) -> str:
    fn = filename.lower()
    if "speicher" in fn:
        return KAT_SPEICHER
    if "wechselrichter" in fn:
        return KAT_WECHSELRICHTER
    return KAT_MODULE


def _make_row(anbieter, kategorie, produkt, preistyp, preis, einheit, datum, quelle):
    return {
        "Anbieter": anbieter,
        "Kategorie": kategorie,
        "Produkt": produkt,
        "Preistyp": preistyp,
        "Preis": preis,
        "Einheit": einheit,
        "Datum": datum,
        "Quelle": quelle,
    }


# ── TWE: Solarmodule ──────────────────────────────────────────────────────────

def _parse_twe_module(pdf, filename: str) -> list[dict]:
    """
    Format (pro Produkt zwei Zeilen):
      Vertex S+ TSM NEG9R.28   0,134 € / Wp   0,139 € / Wp   0,144 € / Wp
      470W Glas/Glas            62,98 €         65,33 €         67,68 €
    """
    rows = []
    date = _detect_date_from_filename(filename)

    SKIP = {
        "PV-Modul Preisliste", "Die Preise", "Mit dieser",
        "Alle Preise", "Irrtümer", "Bitte beachten", "Preise für größere",
        "TYP Preis Container", "(936 Module)", "(962 Module)",
    }

    wp_re = re.compile(r"(\d+[,\.]\d+)\s*€\s*/\s*Wp")
    eur_re = re.compile(r"(\d{1,3}(?:\.\d{3})*(?:,\d+)?)\s*€(?!\s*/)")  # € but not €/Wp

    for page in pdf.pages:
        text = page.extract_text() or ""
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

        i = 0
        while i < len(lines):
            line = lines[i]

            if _is_footer(line) or any(line.startswith(s) for s in SKIP) or any(s in line for s in SKIP):
                i += 1
                continue

            if "/ Wp" in line:
                wp_prices = [_parse_eur(p) for p in wp_re.findall(line)]
                name_part = wp_re.sub("", line).strip()

                # Zweite Zeile: Watt-Info + Gesamtpreise
                total_prices = []
                watt_desc = ""
                if i + 1 < len(lines):
                    nxt = lines[i + 1]
                    if "/ Wp" not in nxt and not _is_footer(nxt):
                        total_prices = [_parse_eur(p) for p in eur_re.findall(nxt)]
                        watt_desc = eur_re.sub("", nxt).replace("€", "").strip()
                        watt_desc = re.sub(r"\s+", " ", watt_desc).strip()
                        if total_prices:
                            i += 1  # zweite Zeile konsumiert

                product = f"{name_part} {watt_desc}".strip() if watt_desc else name_part

                for j, label in enumerate([PREISTYP_CONTAINER, "6 Paletten", "1 Palette"]):
                    if j < len(wp_prices) and wp_prices[j] is not None:
                        rows.append(_make_row(
                            ANBIETER_TWE, KAT_MODULE, product,
                            label, wp_prices[j], "€/Wp", date, filename,
                        ))
                    if j < len(total_prices) and total_prices[j] is not None:
                        rows.append(_make_row(
                            ANBIETER_TWE, KAT_MODULE, product,
                            label, total_prices[j], "€/Modul", date, filename,
                        ))

            i += 1

    return rows


# ── TWE: Speichersysteme ──────────────────────────────────────────────────────

_SECTION_HEADERS_SPEICHER = {
    "Sungrow SBR Serie", "SBH", "EK - Hochvoltbatterien",
    "EQ5000 - Hochvoltbatterien", "BYD Hochvolt HVS/HVM",
    "BYD Hochvolt HVB", "Zubehör", "Speichersysteme",
    "P reisliste", "Inhalt", "Artikelbezeichnung Preis",
    "Artikelbezeichnung",
}

_BRAND_SECTION_KEYWORDS = ["Sungrow", "FoxESS", "SMA", "BYD"]


def _is_speicher_section_header(line: str) -> bool:
    if any(line == h for h in _SECTION_HEADERS_SPEICHER):
        return True
    # Zeilen, die nur aus Herstellerbezeichnungen ohne Preis bestehen
    if not _GERMAN_PRICE_RE.search(line):
        if any(kw in line for kw in ["Hochvolt", "Serie", "Zubehör", "Inhalt", "Seite"]):
            return True
        # Reine Markenzeilen wie "BYD Hochvolt HVS/HVM"
        if re.match(r"^[A-Z][a-zA-Z\s/\-\.]+$", line) and len(line) < 50:
            return True
    return False


def _parse_twe_speicher(pdf, filename: str) -> list[dict]:
    """
    Format:  Artikelbezeichnung  PREIS €
    Einige Einträge zwei-/mehrzeilig:
      Sungrow SBR 3.2kWh
      795 €
      Batteriemodul
    """
    rows = []
    date = _detect_date_from_filename(filename)

    SKIP_LINES = {"Artikelbezeichnung Preis", "Preis", "Artikelbezeichnung"}

    for page in pdf.pages:
        text = page.extract_text() or ""
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

        pending_name: str | None = None

        for line in lines:
            if _is_footer(line) or line in SKIP_LINES:
                pending_name = None
                continue
            if any(kw in line for kw in [
                "Die Preise", "Mit dieser", "Alle Preise", "Irrtümer",
                "Bitte beachten", "Speichersysteme", "P reisliste", "Inhalt",
            ]):
                pending_name = None
                continue

            prices = _all_prices_in_line(line)

            if prices:
                # letzter Preis ist der aktuell gültige (bei SALE)
                price = prices[-1]

                # Produktname: alles vor dem ersten EUR-Betrag
                m = _GERMAN_PRICE_RE.search(line)
                name_part = line[:m.start()].strip() if m else ""

                if name_part:
                    product = name_part
                    pending_name = None
                elif pending_name:
                    product = pending_name
                    pending_name = None
                else:
                    product = "Unbekannt"

                if product and len(product) > 2 and product not in SKIP_LINES:
                    rows.append(_make_row(
                        ANBIETER_TWE, KAT_SPEICHER, product,
                        PREISTYP_STUECK, price, "€", date, filename,
                    ))
            else:
                # Keine Preiszahl → potentieller Produktname (für nächste Zeile)
                if len(line) > 3 and not _is_speicher_section_header(line):
                    pending_name = line
                else:
                    pending_name = None

    return rows


# ── TWE: Wechselrichter ───────────────────────────────────────────────────────

# Zeilen-Muster: "TYP [MPP PHASES WEIGHT] PRICE €"
# Bei Sonderangeboten: "TYP [SPECS] OLDPRICE € NEWPRICE €"
_WECH_PRODUCT_RE = re.compile(
    r"^(.+?)"                                    # Produktname (nicht-gierig)
    r"(?:\s+(\d+)\s+(\d+)\s+([\d,]+))?"          # opt. MPP PHASEN GEWICHT
    r"\s+(\d{1,3}(?:\.\d{3})*(?:,\d+)?)\s*€"    # Preis
    r"(?:\s+(\d{1,3}(?:\.\d{3})*(?:,\d+)?)\s*€)?$"  # opt. Sale-Preis
)

_WECH_SKIP_KEYWORDS = [
    "Die Preise", "Mit dieser", "Alle Preise", "Irrtümer", "Bitte beachten",
    "Wechselrichter", "P reisliste", "Inhalt",
    "TYP MPP TRACKER", "PHASEN", "GEWICHT",
]

_WECH_SECTION_KEYWORDS = [
    "SG", "SH Hybrid", "Zubehör", "Wallbox", "H3-Smart", "H3-Pro",
    "T-G3", "Sunny Tripower", "CORE", "Solar", "Sungrow", "FoxESS", "SMA",
]


def _is_wech_section_header(line: str) -> bool:
    if _GERMAN_PRICE_RE.search(line):
        return False
    if "auf Anfrage" in line and not _GERMAN_PRICE_RE.search(line):
        return False
    # reine Abschnittsüberschriften sind kurz und ohne Preise
    if len(line) < 60 and not any(c.isdigit() for c in line):
        return True
    # bekannte Abschnittstitel
    if line in _WECH_SECTION_KEYWORDS:
        return True
    return False


def _parse_twe_wechselrichter(pdf, filename: str) -> list[dict]:
    """
    Format: TYP [MPP PHASEN GEWICHT] PREIS €  (optional zwei Preise bei Sale)
    Zweizeilig (Wallbox):
      AC22E-01 V112 Wallbox
      22 kW 6 545 €
    """
    rows = []
    date = _detect_date_from_filename(filename)
    pending_name: str | None = None

    for page in pdf.pages:
        text = page.extract_text() or ""
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

        for line in lines:
            if _is_footer(line):
                continue
            if any(kw in line for kw in _WECH_SKIP_KEYWORDS):
                pending_name = None
                continue

            if "auf Anfrage" in line:
                # Preis nicht verfügbar – trotzdem als Eintrag speichern
                name_part = line.replace("auf Anfrage", "").replace("-SALE%", "").strip()
                name_part = re.sub(r"\s+\d+\s+\d+\s+[\d,]+\s*$", "", name_part).strip()
                if pending_name and not name_part:
                    name_part = pending_name
                    pending_name = None
                if name_part and len(name_part) > 2:
                    rows.append(_make_row(
                        ANBIETER_TWE, KAT_WECHSELRICHTER, name_part,
                        PREISTYP_STUECK, None, "€", date, filename,
                    ))
                continue

            prices = _all_prices_in_line(line)

            if prices:
                price = prices[-1]  # bei Sale-Preisen den letzten nehmen

                # Produktname aus der Zeile: vor dem ersten EUR-Preis
                m = _GERMAN_PRICE_RE.search(line)
                raw_name = line[:m.start()].strip() if m else ""

                # MPP / Phasen / Gewicht am Ende des Namens entfernen
                raw_name = re.sub(r"\s+\d+\s+\d+\s+[\d,]+\s*$", "", raw_name).strip()
                # "22 kW 6" am Anfang von Zeilen entfernen
                raw_name = re.sub(r"^[\d\s,kW]+\s*", "", raw_name).strip()
                # -SALE% entfernen
                raw_name = raw_name.replace("-SALE%", "").strip()

                if raw_name:
                    product = raw_name
                    pending_name = None
                elif pending_name:
                    product = pending_name
                    pending_name = None
                else:
                    product = "Unbekannt"

                if product and len(product) > 2:
                    rows.append(_make_row(
                        ANBIETER_TWE, KAT_WECHSELRICHTER, product,
                        PREISTYP_STUECK, price, "€", date, filename,
                    ))
            else:
                if _is_wech_section_header(line):
                    pending_name = None
                elif len(line) > 3:
                    pending_name = line  # potentieller mehrzeiliger Produktname

    return rows


# ── TWE: Dispatcher ───────────────────────────────────────────────────────────

def parse_twe(file_obj, filename: str) -> pd.DataFrame:
    """Erkennt TWE-Listentyp anhand des Dateinamens und parst entsprechend."""
    kat = _detect_twe_type(filename)
    with pdfplumber.open(file_obj) as pdf:
        if kat == KAT_MODULE:
            rows = _parse_twe_module(pdf, filename)
        elif kat == KAT_SPEICHER:
            rows = _parse_twe_speicher(pdf, filename)
        else:
            rows = _parse_twe_wechselrichter(pdf, filename)
    return pd.DataFrame(rows, columns=SCHEMA) if rows else pd.DataFrame(columns=SCHEMA)


# ── Pesasun ───────────────────────────────────────────────────────────────────

def _detect_pesasun_category(text: str) -> str:
    tl = text.lower()
    if "powerocean" in tl:
        return "EcoFlow PowerOcean"
    if "ecoflow zubehör" in tl or "powerinsight" in tl or "smartplug" in tl or "powerpulse" in tl:
        return "EcoFlow Zubehör"
    if "wärmepumpe" in tl or "powerheat" in tl or "powerglow" in tl:
        return "EcoFlow Wärmepumpe"
    if "solplanet wechselrichter" in tl:
        return "Solplanet Wechselrichter"
    if "solplanet" in tl and ("speicher" in tl or "ai-hb" in tl):
        return "Solplanet Speicher"
    if "sigenergy controller" in tl:
        return "Sigenergy Controller"
    if "sigenergy battery" in tl:
        return "Sigenergy Battery"
    if "sigenergy zubehör" in tl or "installation kit" in tl or "sensor three" in tl:
        return "Sigenergy Zubehör"
    return KAT_MODULE


def _parse_pesasun_module_page(lines: list[str], page_no: int, filename: str) -> list[dict]:
    """Seiten mit Container/Palette/Kommission-Preisen."""
    rows = []
    # Preise: eine Zeile mit 2-3 Zahlen (z.B. "62 63,90 65,90")
    price_line = None
    product_lines = []

    for line in lines:
        if line.lower() in ["www.pesasun.com", "www.peeek.com"] or "seite |" in line.lower():
            continue
        if line in ["Container", "Palette", "Kommission"] or line == "Container Palette Kommission":
            continue

        # Preiszeile: nur Zahlen (Komma als Dezimaltrennzeichen) und Leerzeichen
        if re.match(r"^[\d\s,\.]+$", line):
            price_line = line
        else:
            product_lines.append(line)

    if not price_line:
        return rows

    # Preise parsen
    price_values = []
    for tok in re.split(r"\s+", price_line.strip()):
        p = _parse_eur(tok)
        if p is not None:
            price_values.append(p)

    if len(price_values) < 1:
        return rows

    # Produktname: erste zwei produktrelevante Zeilen
    skip_words = ["Container", "Palette", "Kommission", "PESASUN", "www."]
    candidates = [
        ln for ln in product_lines
        if not any(s.lower() in ln.lower() for s in skip_words)
        and not re.fullmatch(r"[\d\s,\.€/]+", ln)
        and len(ln) >= 3
    ]
    product_name = " ".join(candidates[:2]).strip()

    if not product_name:
        return rows

    for i, label in enumerate([PREISTYP_CONTAINER, PREISTYP_PALETTE, "Kommission"]):
        if i < len(price_values):
            rows.append(_make_row(
                ANBIETER_PESASUN, KAT_MODULE, product_name,
                label, price_values[i], "€/Modul", "2026-03", filename,
            ))

    return rows


_PESASUN_SKIP_LINES = {
    "www.pesasun.com", "www.peeek.com", "model", "preis",
    "model preis", "ecoflow powerocean", "ecoflow powerocean batteriespeicher",
    "sigenergy battery", "sigenergy zubehör", "solplanet ai-hb g2 pro",
    "hybrid wechselrichter", "batteriespeicher",
}

_PESASUN_SECTION_RE = re.compile(
    r"^(ecoflow|solplanet|sigenergy|hybrid wechselrichter"
    r"|batteriespeicher|powerheat|powerglow|powerinsight"
    r"|sigen battery|sigenergy controller|sigenergy zubehör)$",
    re.IGNORECASE,
)


def _parse_pesasun_model_price_page(lines: list[str], page_no: int, filename: str) -> list[dict]:
    """
    Seiten mit Model-/Preis-Spalten (EcoFlow, Solplanet, Sigenergy).
    Unterstützt kompakte Preise (1230€), mehrzeilige Produktnamen und
    mehrere Kategorieabschnitte innerhalb einer Seite.
    """
    rows = []
    full_text = " ".join(lines)
    # Startkategorie aus Gesamttext; wird inline aktualisiert
    current_kat = _detect_pesasun_category(full_text)

    pending_name: str | None = None

    for line in lines:
        ll = line.lower().strip()

        # Irrelevante Zeilen überspringen
        if ll in _PESASUN_SKIP_LINES:
            continue
        if "seite |" in ll or "peeek" in ll:
            continue
        if _PESASUN_SECTION_RE.match(ll):
            # Kategorie bei Abschnittswechsel aktualisieren
            current_kat = _detect_pesasun_category(line)
            pending_name = None
            continue
        if ll in ("pesasun gmbh",):
            continue

        # Preis in der Zeile suchen (inklusive 4-stelliger Kompaktpreise)
        pm = _GERMAN_PRICE_RE.search(line)

        if pm:
            price = _parse_eur(pm.group(1))
            name_part = line[:pm.start()].strip()

            if name_part:
                product = name_part
                pending_name = None
            elif pending_name:
                product = pending_name
                pending_name = None
            else:
                product = None

            if product and len(product) > 2:
                rows.append(_make_row(
                    ANBIETER_PESASUN, current_kat, product,
                    PREISTYP_STUECK, price, "€", "2026-03", filename,
                ))
        else:
            # Kein Preis → potenzieller Produktname (für mehrzeilige Einträge)
            if len(line) > 2:
                pending_name = (
                    f"{pending_name} {line}".strip() if pending_name else line
                )
            else:
                pending_name = None

    return rows


def parse_pesasun(file_obj, filename: str) -> pd.DataFrame:
    rows = []
    with pdfplumber.open(file_obj) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

            if (
                "Container" in text
                and "Palette" in text
                and "Kommission" in text
                and "Model" not in text
            ):
                rows.extend(_parse_pesasun_module_page(lines, page_no, filename))
            elif "Model" in text and "Preis" in text:
                rows.extend(_parse_pesasun_model_price_page(lines, page_no, filename))
            # Seiten ohne klare Struktur (Deckblatt, Intro) überspringen

    return pd.DataFrame(rows, columns=SCHEMA) if rows else pd.DataFrame(columns=SCHEMA)


# ── pv partners ───────────────────────────────────────────────────────────────

_PVP_FOOTER = [
    "pv partners AG", "Großbeerenstr", "Telefon", "Telefax",
    "www.pv-partners.de", "info@pv-partners.de",
]

_PVP_SKIP_KEYWORDS = [
    "PV-Module:", "Gültig für", "Module (Kommissioniert",
    "Module (Projektmengen", "Verfügbarkeit auf Anfrage",
    "Mindermengenzuschlag:", "bei Abnahme", "Legende:", "Die Verfügbarkeit",
    "Alle Angaben", "Produktgarantien", "Alle Preise",
    "ab 1 Palette", "ab 10 Paletten", "ab 1 Container",
    "≤ 50 kWp", "> 50 kWp", "Preis je Wp", "Verfügbarkeit",
    "Sonderpreisliste", "Wechselrichter & Zubehör",
    "Einzelabnahme", "Abnahme >", "Preis Preis",
    "Solange der Vorat", "Solange der Vorrat",
]

# Brands listed in pv partners documents
_PVP_BRANDS = [
    "DMEGC Solar", "JA Solar", "LONGI", "Solyco",
    "Trina Solar", "Fronius", "GoodWe", "Huawei", "KOSTAL", "SMA",
]


def _detect_date_pvpartners(text: str, filename: str) -> str:
    """'Gültig für Bestellungen ab 11.05.2026' oder '_260511.pdf' → '2026-05-11'."""
    m = re.search(r"Gültig.*ab\s+(\d{2})\.(\d{2})\.(\d{4})", text)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    m = re.search(r"_(\d{2})(\d{2})(\d{2})\.pdf$", filename, re.IGNORECASE)
    if m:
        return f"20{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return "unbekannt"


def _normalize_pvp_line(line: str) -> str:
    """Fix PDF artefact '0 ,133 €' → '0,133 €'."""
    return re.sub(r"(\d)\s+,(\d)", r"\1,\2", line)


def _extract_pvp_product_name(before_price: str) -> str:
    """
    Extract 'MODEL [WATT]Wp' from the pre-price text of a pv partners line.
    Handles three formats:
      • single-token model + watt:   'DM465G12RT-B48HBW 465 Wp …'
      • two-token model + watt:      'R-WG 96n.5/465 465 Wp …'
      • multi-token model, no watt:  'Hi-Mo X10 LR7-54HVH-490M …'
    """
    # single token + WATT Wp
    m = re.match(r"^(\S+)\s+(\d+)\s+Wp\b", before_price)
    if m:
        return f"{m.group(1)} {m.group(2)}Wp"
    # two tokens + WATT Wp  (e.g. Solyco: "R-WG 96n.5/465 465 Wp")
    m = re.match(r"^(\S+\s+\S+)\s+(\d+)\s+Wp\b", before_price)
    if m:
        return f"{m.group(1)} {m.group(2)}Wp"
    # fallback: first three whitespace-separated tokens (LONGI style)
    tokens = before_price.split()
    return " ".join(tokens[:3]) if tokens else "Unbekannt"


def _is_pvp_brand_header(line: str) -> bool:
    """True when the line is a brand/section header with no EUR price."""
    if _GERMAN_PRICE_RE.search(line):
        return False
    return any(b.lower() in line.lower() for b in _PVP_BRANDS)


def _parse_pvpartners_module(pdf, filename: str) -> list[dict]:
    """
    pv partners Modulpreisliste:
    • Abschnitt 1 'Kommissioniert & Palettenmengen': 1 Preis/Wp, Preistyp='1 Palette'
    • Abschnitt 2 'Projektmengen': bis zu 2 Preise/Wp,
        1. Preis = '10 Paletten', 2. Preis = 'Container'
    """
    rows = []
    text_all = "\n".join(p.extract_text() or "" for p in pdf.pages)
    date = _detect_date_pvpartners(text_all, filename)
    section = "palette"

    for page in pdf.pages:
        text = page.extract_text() or ""
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

        for line in lines:
            # Section detection
            if "Kommissioniert" in line or "Palettenmengen" in line:
                section = "palette"
                continue
            if "Projektmengen" in line:
                section = "projekt"
                continue

            # Skip footer and irrelevant lines
            if any(kw in line for kw in _PVP_FOOTER + _PVP_SKIP_KEYWORDS):
                continue

            # Normalize space-in-decimal artefact
            line_n = _normalize_pvp_line(line)

            # Brand/section headers have no price → skip
            if _is_pvp_brand_header(line_n):
                continue

            prices = _all_prices_in_line(line_n)
            if not prices:
                continue

            # Everything before the first price
            pm = _GERMAN_PRICE_RE.search(line_n)
            before_price = line_n[:pm.start()].strip() if pm else ""

            # Remove availability marker at the end of before_price (section 1 only)
            before_price = re.sub(
                r"\s*(verfügbar|KW\s*\d+/\d+|auf\s*Anfrage)\s*$",
                "", before_price, flags=re.IGNORECASE,
            ).strip()

            product = _extract_pvp_product_name(before_price)

            if not product or len(product) < 3:
                continue
            # Skip footnote/surcharge lines with zero price
            if prices[0] == 0.0:
                continue

            if section == "palette":
                rows.append(_make_row(
                    ANBIETER_PVPARTNERS, KAT_MODULE, product,
                    "1 Palette", prices[0], "€/Wp", date, filename,
                ))
            else:
                rows.append(_make_row(
                    ANBIETER_PVPARTNERS, KAT_MODULE, product,
                    "10 Paletten", prices[0], "€/Wp", date, filename,
                ))
                if len(prices) >= 2:
                    rows.append(_make_row(
                        ANBIETER_PVPARTNERS, KAT_MODULE, product,
                        PREISTYP_CONTAINER, prices[1], "€/Wp", date, filename,
                    ))

    return rows


def _parse_pvpartners_sonderliste(pdf, filename: str) -> list[dict]:
    """
    pv partners Sonderpreisliste (Wechselrichter & Zubehör):
    Zwei optionale Preisspalten: Einzelabnahme / Abnahme > 3 Stk.
    """
    rows = []
    text_all = "\n".join(p.extract_text() or "" for p in pdf.pages)
    date = _detect_date_pvpartners(text_all, filename)

    for page in pdf.pages:
        text = page.extract_text() or ""
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

        for line in lines:
            if any(kw in line for kw in _PVP_FOOTER + _PVP_SKIP_KEYWORDS):
                continue

            line_n = _normalize_pvp_line(line)
            prices = _all_prices_in_line(line_n)
            if not prices:
                continue

            pm = _GERMAN_PRICE_RE.search(line_n)
            product = line_n[:pm.start()].strip() if pm else ""
            if not product or len(product) < 3:
                continue

            # Determine category from product name
            kat = (
                "Zubehör"
                if any(kw in product for kw in ["Meter", "Smart Meter", "Zubehör", "Sensor"])
                else KAT_WECHSELRICHTER
            )

            rows.append(_make_row(
                ANBIETER_PVPARTNERS, kat, product,
                "Einzelabnahme", prices[0], "€", date, filename,
            ))
            if len(prices) >= 2:
                rows.append(_make_row(
                    ANBIETER_PVPARTNERS, kat, product,
                    "Mengenpreis (>3 Stk.)", prices[1], "€", date, filename,
                ))

    return rows


def parse_pvpartners(file_obj, filename: str) -> pd.DataFrame:
    """Erkennt pv-partners-Listentyp aus Dateiinhalt und parst entsprechend."""
    with pdfplumber.open(file_obj) as pdf:
        first_page_text = pdf.pages[0].extract_text() or ""
        if "PV-Module" in first_page_text or "Modulpreis" in filename.lower():
            rows = _parse_pvpartners_module(pdf, filename)
        else:
            rows = _parse_pvpartners_sonderliste(pdf, filename)
    return pd.DataFrame(rows, columns=SCHEMA) if rows else pd.DataFrame(columns=SCHEMA)


# ── Ökoteam Solar ─────────────────────────────────────────────────────────────
#
# Format (pro Zeile):  NAME  ART_NUM(5-stellig)  HUAWEI_NUM  PREIS
# Preisformat:         "1 088,00"  (Leerzeichen als Tausendertrenner, Komma als Dezimal)
# Kein €-Zeichen.

# Preis am Zeilenende: 1–4 Zifferngruppen à 3 Ziffern, durch Leerzeichen getrennt,
# gefolgt von Komma und zwei Dezimalstellen.
# Das führende \s stellt sicher, dass der Preis nicht mitten in einer Artikelnummer
# gematcht wird (z. B. "016-006 621,00" → nur "621,00", nicht "006 621,00").
_OKT_PRICE_RE = re.compile(r"\s(\d{1,3}(?:\s\d{3})*,\d{2})\s*$")

# Produktname = alles vor der 5-stelligen Artikelnummer (die immer von Leerzeichen umgeben ist)
_OKT_ARTNUM_RE = re.compile(r"^(.*?)\s+(\d{5})\s+")

_OKT_SKIP = [
    "Hersteller", "Art-Num", "Huawei Bezeichnung",
    "Irrtümer", "Schreibfehler", "typographical",
    "voraussichtlich verfügbar",
]


def _detect_date_oekoteam(filename: str) -> str:
    """'Preisliste_..._03052026.pdf'  →  '2026-05-03'  (DDMMYYYY am Dateiende)."""
    m = re.search(r"(\d{2})(\d{2})(\d{4})\.pdf$", filename, re.IGNORECASE)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return "unbekannt"


def _detect_oekoteam_category(product: str) -> str:
    pl = product.lower()
    if "luna2000" in pl:
        return KAT_SPEICHER
    if "leistungsoptimierer" in pl:
        return "Zubehör"
    if "sun2000" in pl or "sun5000" in pl:
        return KAT_WECHSELRICHTER
    return "Zubehör"


def _parse_okt_price(price_str: str) -> float | None:
    """'1 088,00'  →  1088.0  (Leerzeichen-Tausender + Komma-Dezimal, kein €)."""
    s = price_str.strip().replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_oekoteam(file_obj, filename: str) -> pd.DataFrame:
    """
    Parser für Ökoteam Solar Preislisten.
    Jede Produktzeile: NAME  5-DIGIT-ARTNUM  HUAWEI-ARTNUM  PREIS
    """
    rows = []
    date = _detect_date_oekoteam(filename)

    with pdfplumber.open(file_obj) as pdf:
        for page in pdf.pages:
            lines = [ln.strip() for ln in (page.extract_text() or "").splitlines()
                     if ln.strip()]

            for line in lines:
                # Seitenangaben und bekannte Nicht-Produkt-Zeilen überspringen
                if line.isdigit():
                    continue
                if any(kw in line for kw in _OKT_SKIP):
                    continue

                # Preis am Zeilenende extrahieren
                price_m = _OKT_PRICE_RE.search(line)
                if not price_m:
                    continue

                price = _parse_okt_price(price_m.group(1))
                if price is None or price <= 0:
                    continue

                # Alles vor dem Preis
                before_price = line[:price_m.start()].strip()

                # Produktname = alles vor der 5-stelligen Artikelnummer
                name_m = _OKT_ARTNUM_RE.match(before_price)
                product = name_m.group(1).strip() if name_m else before_price

                if not product or len(product) < 3:
                    continue

                kategorie = _detect_oekoteam_category(product)

                rows.append(_make_row(
                    ANBIETER_OEKOTEAM, kategorie, product,
                    PREISTYP_STUECK, price, "€", date, filename,
                ))

    return pd.DataFrame(rows, columns=SCHEMA) if rows else pd.DataFrame(columns=SCHEMA)


# ── CSV / JSON ────────────────────────────────────────────────────────────────

def parse_csv(file_obj, filename: str) -> pd.DataFrame:
    """Generischer CSV-Import; erwartet Spalten aus SCHEMA oder kompatibel."""
    df = pd.read_csv(file_obj)
    for col in SCHEMA:
        if col not in df.columns:
            df[col] = None
    return df[SCHEMA]


def parse_json(file_obj, filename: str) -> pd.DataFrame:
    """Generischer JSON-Import (Liste von Objekten)."""
    df = pd.read_json(file_obj)
    for col in SCHEMA:
        if col not in df.columns:
            df[col] = None
    return df[SCHEMA]


# ── Datei-Dispatcher ──────────────────────────────────────────────────────────

def parse_file(file_obj, filename: str) -> pd.DataFrame:
    fn_lower = filename.lower()
    if fn_lower.endswith(".csv"):
        return parse_csv(file_obj, filename)
    if fn_lower.endswith(".json"):
        return parse_json(file_obj, filename)
    # PDF
    if "twe" in fn_lower:
        return parse_twe(file_obj, filename)
    if "pesasun" in fn_lower or "produktkatalog" in fn_lower:
        return parse_pesasun(file_obj, filename)
    if "pv_partners" in fn_lower or "pv-partners" in fn_lower:
        return parse_pvpartners(file_obj, filename)
    if "oekoteam" in fn_lower or "ökoteam" in fn_lower or "preisliste_huawei" in fn_lower:
        return parse_oekoteam(file_obj, filename)
    # Fallback: als TWE versuchen
    return parse_twe(file_obj, filename)


def load_from_directory(directory: Path) -> pd.DataFrame:
    frames = []
    supported = (".pdf", ".csv", ".json")
    for path in sorted(directory.rglob("*")):
        if path.suffix.lower() in supported and not path.name.startswith("."):
            try:
                with open(path, "rb") as f:
                    df = parse_file(f, path.name)
                if not df.empty:
                    frames.append(df)
            except Exception as e:
                st.warning(f"Fehler beim Laden von {path.name}: {e}")
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=SCHEMA)


# ── XLSX-Export ───────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL = PatternFill("solid", fgColor="D6E4F0")


def _to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Preisvergleich"

    headers = list(df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = _ALT_FILL

    # Spaltenbreite automatisch
    for col_idx, col in enumerate(df.columns, start=1):
        lens = [len(str(col))] + [len(str(v)) for v in df[col] if v is not None]
        max_len = max(lens) if lens else 10
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Streamlit App ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="PV Preisvergleich",
    page_icon="☀️",
    layout="wide",
)

# ── Auth-Hilfsfunktionen ──────────────────────────────────────────────────────

AUTH_CONFIG_PATH = Path("auth_config.yaml")
ROLES = ["admin", "viewer"]


def _load_auth_config() -> dict:
    """Lädt auth_config.yaml. Erstellt Standardkonfiguration falls nicht vorhanden."""
    if AUTH_CONFIG_PATH.exists():
        with open(AUTH_CONFIG_PATH, "r") as f:
            return yaml.safe_load(f)
    # Erstmalige Einrichtung: Standard-Admin anlegen
    default_pw = Hasher.hash("admin")
    cfg = {
        "credentials": {
            "usernames": {
                "admin": {
                    "name": "Administrator",
                    "email": "admin@local",
                    "password": default_pw,
                    "roles": ["admin"],
                }
            }
        },
        "cookie": {
            "name": "pv_preisvergleich",
            "key": secrets.token_hex(32),
            "expiry_days": 30,
        },
    }
    _save_auth_config(cfg)
    return cfg


def _save_auth_config(cfg: dict) -> None:
    with open(AUTH_CONFIG_PATH, "w") as f:
        yaml.dump(cfg, f, allow_unicode=True, default_flow_style=False)


def _current_roles() -> list[str]:
    """Gibt die Rollen des eingeloggten Benutzers zurück."""
    username = st.session_state.get("username", "")
    cfg = st.session_state.get("_auth_cfg", {})
    user = cfg.get("credentials", {}).get("usernames", {}).get(username, {})
    return user.get("roles", [])


def _is_admin() -> bool:
    return "admin" in _current_roles()


# ── Auth initialisieren ───────────────────────────────────────────────────────

cfg = _load_auth_config()
st.session_state["_auth_cfg"] = cfg

authenticator = stauth.Authenticate(
    credentials=cfg["credentials"],
    cookie_name=cfg["cookie"]["name"],
    cookie_key=cfg["cookie"]["key"],
    cookie_expiry_days=cfg["cookie"]["expiry_days"],
    auto_hash=False,
)

# ── Login-Seite ───────────────────────────────────────────────────────────────

authenticator.login(
    location="main",
    fields={"Form name": "☀️ PV-Preisvergleich · Anmeldung",
            "Username": "Benutzername",
            "Password": "Passwort",
            "Login": "Anmelden"},
)

auth_status = st.session_state.get("authentication_status")
username    = st.session_state.get("username", "")
name        = st.session_state.get("name", "")

if auth_status is False:
    st.error("Benutzername oder Passwort falsch.")
    st.stop()

if auth_status is None:
    st.info("Bitte melden Sie sich an.")
    st.stop()

# ─── ab hier: Benutzer ist authentifiziert ───────────────────────────────────

# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown(f"**{name}** `{' · '.join(_current_roles())}`")
    authenticator.logout(
        button_name="🚪 Abmelden",
        location="sidebar",
        key="sidebar_logout",
    )
    st.divider()

    st.header("Datenquellen")

    lists_dir = Path("lists")
    auto_load = st.checkbox(
        "Automatisch aus `lists/` laden",
        value=lists_dir.exists(),
        disabled=not lists_dir.exists(),
    )

    uploaded = st.file_uploader(
        "Weitere Dateien hochladen (PDF, CSV, JSON)",
        type=["pdf", "csv", "json"],
        accept_multiple_files=True,
    )

    if st.button("🔄 Neu laden", use_container_width=True):
        st.cache_data.clear()

    st.divider()
    st.header("Filter")

# ── Daten laden ────────────────────────────────────────────────────────────────

st.title("☀️ PV-Preisvergleich")
st.caption("Lieferanten: TWE Solar GmbH · Pesasun GmbH · pv partners AG · Ökoteam Solar")


@st.cache_data(show_spinner="Preislisten werden geparst…")
def _load_directory(directory: str) -> pd.DataFrame:
    return load_from_directory(Path(directory))


frames: list[pd.DataFrame] = []

if auto_load and lists_dir.exists():
    df_dir = _load_directory(str(lists_dir))
    if not df_dir.empty:
        frames.append(df_dir)

for uf in uploaded or []:
    try:
        df_up = parse_file(uf, uf.name)
        if not df_up.empty:
            frames.append(df_up)
    except Exception as e:
        st.sidebar.error(f"{uf.name}: {e}")

if frames:
    df_all = pd.concat(frames, ignore_index=True)
    df_all["Preis"] = pd.to_numeric(df_all["Preis"], errors="coerce")
    df_all = df_all.drop_duplicates(
        subset=["Anbieter", "Produkt", "Preistyp", "Datum", "Einheit"]
    )
else:
    df_all = pd.DataFrame(columns=SCHEMA)

# ── Sidebar-Filter ─────────────────────────────────────────────────────────────

with st.sidebar:
    if df_all.empty:
        st.info("Noch keine Daten geladen.")
    else:
        anbieter_opts = sorted(df_all["Anbieter"].dropna().unique())
        kat_opts      = sorted(df_all["Kategorie"].dropna().unique())
        datum_opts    = sorted(df_all["Datum"].dropna().unique(), reverse=True)
        einheit_opts  = sorted(df_all["Einheit"].dropna().unique())

        sel_anbieter = st.multiselect("Anbieter",           anbieter_opts, default=anbieter_opts)
        sel_kat      = st.multiselect("Kategorie",          kat_opts,      default=kat_opts)
        sel_datum    = st.multiselect("Preislisten-Datum",  datum_opts,    default=datum_opts)
        sel_einheit  = st.multiselect("Einheit",            einheit_opts,  default=einheit_opts)
        search       = st.text_input("Produkt suchen", placeholder="z.B. Sungrow, 470W …")

        mask = (
            df_all["Anbieter"].isin(sel_anbieter)
            & df_all["Kategorie"].isin(sel_kat)
            & df_all["Datum"].isin(sel_datum)
            & df_all["Einheit"].isin(sel_einheit)
        )
        if search:
            mask &= df_all["Produkt"].str.contains(search, case=False, na=False)

        df_filtered = df_all[mask].copy()

# ── Hauptbereich ───────────────────────────────────────────────────────────────

if df_all.empty:
    st.info(
        "Keine Daten geladen. Aktiviere das automatische Laden aus dem `lists/`-Verzeichnis "
        "oder lade Dateien über die Seitenleiste hoch."
    )
    st.stop()

# Kennzahlen
col1, col2, col3, col4 = st.columns(4)
col1.metric("Einträge gesamt",  f"{len(df_filtered):,}")
col2.metric("Produkte",         df_filtered["Produkt"].nunique())
col3.metric("Anbieter",         df_filtered["Anbieter"].nunique())
col4.metric("Preislisten-Daten", df_filtered["Datum"].nunique())

st.divider()

# Tabs – Admins sehen zusätzlich den Benutzerverwaltungs-Tab
_tab_labels = ["📋 Alle Preise", "⚖️ Preisvergleich", "📈 Preishistorie", "💾 Export"]
if _is_admin():
    _tab_labels.append("👤 Benutzerverwaltung")

_tabs = st.tabs(_tab_labels)
tab_all, tab_compare, tab_history, tab_export = _tabs[:4]
tab_users = _tabs[4] if _is_admin() else None

# ── Tab: Alle Preise ──────────────────────────────────────────────────────────

with tab_all:
    st.subheader("Alle Preise")

    sort_col = st.selectbox(
        "Sortieren nach", ["Kategorie", "Anbieter", "Produkt", "Preis", "Datum"],
        index=0, key="sort_all",
    )
    df_show = df_filtered.sort_values([sort_col, "Produkt"]).reset_index(drop=True)

    st.dataframe(
        df_show.style.format({"Preis": lambda x: f"{x:,.2f}" if pd.notna(x) else "auf Anfrage"}),
        use_container_width=True,
        height=520,
    )

# ── Tab: Preisvergleich ───────────────────────────────────────────────────────

with tab_compare:
    st.subheader("Preisvergleich zwischen Anbietern")

    kat_vergleich = st.selectbox(
        "Kategorie", sorted(df_filtered["Kategorie"].dropna().unique()), key="kat_vergleich"
    )
    einheit_vergleich = st.selectbox(
        "Einheit", sorted(df_filtered["Einheit"].dropna().unique()), key="ein_vergleich"
    )
    preistyp_vergleich = st.selectbox(
        "Preistyp",
        sorted(df_filtered["Preistyp"].dropna().unique()),
        key="pt_vergleich",
    )

    df_v = df_filtered[
        (df_filtered["Kategorie"] == kat_vergleich)
        & (df_filtered["Einheit"] == einheit_vergleich)
        & (df_filtered["Preistyp"] == preistyp_vergleich)
    ].copy()

    if df_v.empty:
        st.info("Keine Daten für diese Kombination.")
    else:
        df_v["Datum"] = df_v["Datum"].astype(str)
        latest   = df_v.groupby("Anbieter")["Datum"].max().reset_index(name="MaxDatum")
        df_v     = df_v.merge(latest, on="Anbieter")
        df_latest = df_v[df_v["Datum"] == df_v["MaxDatum"]].drop(columns=["MaxDatum"])

        pivot = (
            df_latest
            .pivot_table(index="Produkt", columns="Anbieter", values="Preis", aggfunc="min")
            .reset_index()
        )

        anbieter_cols = [c for c in pivot.columns if c != "Produkt"]
        if len(anbieter_cols) >= 2:
            a, b = anbieter_cols[0], anbieter_cols[1]
            pivot["Differenz"] = pivot[a] - pivot[b]
            pivot["Günstiger"] = pivot.apply(
                lambda r: a if pd.notna(r[a]) and (pd.isna(r[b]) or r[a] <= r[b]) else b,
                axis=1,
            )

        fmt = {col: "{:,.2f}" for col in anbieter_cols}
        if "Differenz" in pivot.columns:
            fmt["Differenz"] = "{:+,.2f}"

        st.dataframe(pivot.style.format(fmt, na_rep="–"), use_container_width=True, height=480)

        if len(anbieter_cols) >= 1:
            min_price = df_latest.loc[
                df_latest.groupby("Produkt")["Preis"].idxmin()
            ][["Produkt", "Anbieter", "Preis"]].rename(
                columns={"Anbieter": "Günstigster Anbieter", "Preis": "Günstigster Preis"}
            )
            st.markdown("**Günstigster Anbieter je Produkt**")
            st.dataframe(
                min_price.style.format({"Günstigster Preis": "{:,.2f}"}),
                use_container_width=True, height=300,
            )

# ── Tab: Preishistorie ────────────────────────────────────────────────────────

with tab_history:
    st.subheader("Preisentwicklung über Zeit")

    if df_filtered["Datum"].nunique() < 2:
        st.info("Für eine Preisentwicklungsanalyse werden mindestens zwei Preislisten-Daten benötigt.")
    else:
        col_h1, col_h2, col_h3 = st.columns(3)
        with col_h1:
            kat_hist = st.selectbox(
                "Kategorie", sorted(df_filtered["Kategorie"].dropna().unique()), key="kat_hist"
            )
        with col_h2:
            ein_hist = st.selectbox(
                "Einheit", sorted(df_filtered["Einheit"].dropna().unique()), key="ein_hist"
            )
        with col_h3:
            pt_hist = st.selectbox(
                "Preistyp", sorted(df_filtered["Preistyp"].dropna().unique()), key="pt_hist"
            )

        df_h = df_filtered[
            (df_filtered["Kategorie"] == kat_hist)
            & (df_filtered["Einheit"] == ein_hist)
            & (df_filtered["Preistyp"] == pt_hist)
        ].copy()

        if df_h.empty:
            st.info("Keine Daten.")
        else:
            prod_counts = df_h.groupby("Produkt")["Datum"].nunique()
            multi_prods = prod_counts[prod_counts >= 2].index.tolist()

            if not multi_prods:
                st.info("Keine Produkte mit mehreren Preisdaten gefunden.")
            else:
                sel_prod  = st.multiselect(
                    "Produkte auswählen", multi_prods, default=multi_prods[:5], key="hist_prod"
                )
                df_h_sel  = df_h[df_h["Produkt"].isin(sel_prod)]
                pivot_hist = df_h_sel.pivot_table(
                    index="Datum", columns="Produkt", values="Preis", aggfunc="min"
                ).sort_index()
                st.line_chart(pivot_hist, height=400)

                st.markdown("**Preisveränderung (ältestes → neuestes Datum)**")
                change_rows = []
                for prod in sel_prod:
                    sub = df_h_sel[df_h_sel["Produkt"] == prod].sort_values("Datum")
                    if len(sub) >= 2:
                        old_p, new_p = sub.iloc[0]["Preis"], sub.iloc[-1]["Preis"]
                        old_d, new_d = sub.iloc[0]["Datum"],  sub.iloc[-1]["Datum"]
                        if pd.notna(old_p) and pd.notna(new_p) and old_p != 0:
                            change_rows.append({
                                "Produkt": prod,
                                f"Preis {old_d}": old_p,
                                f"Preis {new_d}": new_p,
                                "Veränderung %": (new_p - old_p) / old_p * 100,
                            })
                if change_rows:
                    df_change  = pd.DataFrame(change_rows)
                    num_cols   = [c for c in df_change.columns if c != "Produkt"]
                    st.dataframe(
                        df_change.style
                        .format({c: "{:,.2f}"  for c in num_cols if "%" not in c})
                        .format({c: "{:+.1f}%" for c in num_cols if "%" in c})
                        .background_gradient(subset=["Veränderung %"], cmap="RdYlGn_r"),
                        use_container_width=True,
                    )

# ── Tab: Export ────────────────────────────────────────────────────────────────

with tab_export:
    st.subheader("Daten exportieren")

    st.markdown(
        f"Aktuell gefilterte Datensätze: **{len(df_filtered):,}** "
        f"(von {len(df_all):,} gesamt)"
    )

    col_e1, col_e2 = st.columns(2)

    with col_e1:
        st.download_button(
            label="⬇️ Als XLSX exportieren (gefiltert)",
            data=_to_xlsx_bytes(df_filtered),
            file_name="preisvergleich_gefiltert.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col_e2:
        st.download_button(
            label="⬇️ Als XLSX exportieren (alle Daten)",
            data=_to_xlsx_bytes(df_all),
            file_name="preisvergleich_alle.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.markdown("---")
    st.markdown("**Vorschau der exportierten Daten:**")
    st.dataframe(df_filtered.head(20), use_container_width=True)

# ── Tab: Benutzerverwaltung (nur Admins) ──────────────────────────────────────

if _is_admin() and tab_users is not None:
    with tab_users:
        st.subheader("👤 Benutzerverwaltung")

        # Aktuellen Config neu laden (andere Admins könnten ihn verändert haben)
        live_cfg = _load_auth_config()
        users    = live_cfg["credentials"]["usernames"]

        # ── Benutzertabelle ───────────────────────────────────────────────────
        st.markdown("#### Aktuelle Benutzer")
        user_rows = [
            {
                "Benutzername": uname,
                "Name":         udata.get("name", ""),
                "E-Mail":       udata.get("email", ""),
                "Rollen":       ", ".join(udata.get("roles", [])),
            }
            for uname, udata in sorted(users.items())
        ]
        st.dataframe(pd.DataFrame(user_rows), use_container_width=True, hide_index=True)

        st.divider()

        # ── Aktionen in Spalten ───────────────────────────────────────────────
        col_add, col_edit, col_del = st.columns(3)

        # ── Benutzer hinzufügen ───────────────────────────────────────────────
        with col_add:
            st.markdown("#### ➕ Neuer Benutzer")
            with st.form("form_add_user", clear_on_submit=True):
                new_uname  = st.text_input("Benutzername")
                new_name   = st.text_input("Anzeigename")
                new_email  = st.text_input("E-Mail")
                new_pw     = st.text_input("Passwort", type="password")
                new_pw2    = st.text_input("Passwort wiederholen", type="password")
                new_roles  = st.multiselect("Rollen", ROLES, default=["viewer"])
                submitted  = st.form_submit_button("Benutzer anlegen", use_container_width=True)

            if submitted:
                err = None
                if not new_uname:
                    err = "Benutzername darf nicht leer sein."
                elif new_uname in users:
                    err = f"Benutzername '{new_uname}' existiert bereits."
                elif len(new_pw) < 6:
                    err = "Passwort muss mindestens 6 Zeichen haben."
                elif new_pw != new_pw2:
                    err = "Passwörter stimmen nicht überein."
                elif not new_roles:
                    err = "Mindestens eine Rolle muss gewählt werden."
                if err:
                    st.error(err)
                else:
                    live_cfg["credentials"]["usernames"][new_uname] = {
                        "name":     new_name,
                        "email":    new_email,
                        "password": Hasher.hash(new_pw),
                        "roles":    new_roles,
                    }
                    _save_auth_config(live_cfg)
                    st.success(f"Benutzer **{new_uname}** wurde angelegt.")
                    st.rerun()

        # ── Benutzer bearbeiten ───────────────────────────────────────────────
        with col_edit:
            st.markdown("#### ✏️ Benutzer bearbeiten")
            edit_uname = st.selectbox(
                "Benutzer auswählen", sorted(users.keys()), key="edit_select"
            )
            edit_data  = users.get(edit_uname, {})

            with st.form("form_edit_user"):
                edit_name  = st.text_input("Anzeigename",  value=edit_data.get("name", ""))
                edit_email = st.text_input("E-Mail",        value=edit_data.get("email", ""))
                edit_roles = st.multiselect(
                    "Rollen", ROLES,
                    default=[r for r in edit_data.get("roles", []) if r in ROLES],
                )
                st.markdown("**Passwort ändern** *(leer lassen = unverändert)*")
                edit_pw    = st.text_input("Neues Passwort",            type="password")
                edit_pw2   = st.text_input("Neues Passwort wiederholen", type="password")
                save_edit  = st.form_submit_button("Speichern", use_container_width=True)

            if save_edit:
                err = None
                if not edit_roles:
                    err = "Mindestens eine Rolle muss gewählt werden."
                elif edit_uname == username and "admin" not in edit_roles:
                    err = "Du kannst dir selbst die Admin-Rolle nicht entziehen."
                elif edit_pw and len(edit_pw) < 6:
                    err = "Neues Passwort muss mindestens 6 Zeichen haben."
                elif edit_pw and edit_pw != edit_pw2:
                    err = "Passwörter stimmen nicht überein."
                if err:
                    st.error(err)
                else:
                    live_cfg["credentials"]["usernames"][edit_uname]["name"]  = edit_name
                    live_cfg["credentials"]["usernames"][edit_uname]["email"] = edit_email
                    live_cfg["credentials"]["usernames"][edit_uname]["roles"] = edit_roles
                    if edit_pw:
                        live_cfg["credentials"]["usernames"][edit_uname]["password"] = (
                            Hasher.hash(edit_pw)
                        )
                    _save_auth_config(live_cfg)
                    st.success(f"Benutzer **{edit_uname}** wurde aktualisiert.")
                    st.rerun()

        # ── Benutzer löschen ──────────────────────────────────────────────────
        with col_del:
            st.markdown("#### 🗑️ Benutzer löschen")
            del_candidates = [u for u in sorted(users.keys()) if u != username]
            if not del_candidates:
                st.info("Keine anderen Benutzer vorhanden.")
            else:
                del_uname = st.selectbox(
                    "Benutzer auswählen", del_candidates, key="del_select"
                )
                st.warning(
                    f"Benutzer **{del_uname}** wird dauerhaft gelöscht.",
                    icon="⚠️",
                )
                if st.button("🗑️ Jetzt löschen", type="primary", use_container_width=True):
                    del live_cfg["credentials"]["usernames"][del_uname]
                    _save_auth_config(live_cfg)
                    st.success(f"Benutzer **{del_uname}** wurde gelöscht.")
                    st.rerun()

        # ── Eigenes Passwort ändern ───────────────────────────────────────────
        st.divider()
        st.markdown("#### 🔑 Eigenes Passwort ändern")
        with st.form("form_own_pw", clear_on_submit=True):
            own_pw_new  = st.text_input("Neues Passwort",             type="password")
            own_pw_new2 = st.text_input("Neues Passwort wiederholen", type="password")
            own_pw_save = st.form_submit_button("Passwort ändern", use_container_width=True)

        if own_pw_save:
            if len(own_pw_new) < 6:
                st.error("Passwort muss mindestens 6 Zeichen haben.")
            elif own_pw_new != own_pw_new2:
                st.error("Passwörter stimmen nicht überein.")
            else:
                live_cfg["credentials"]["usernames"][username]["password"] = (
                    Hasher.hash(own_pw_new)
                )
                _save_auth_config(live_cfg)
                st.success("Passwort wurde geändert.")
